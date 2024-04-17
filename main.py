from openpyxl import Workbook, load_workbook
import pandas
import oracledb
from configparser import ConfigParser
from datetime import datetime, timedelta
import dateutil.relativedelta


workbook = Workbook()
trxWorksheet = workbook.create_sheet("Trx")
# workbook.save("output/test.xlsx")

config = ConfigParser()
config.read("config.ini")
USERNAME = config['DATABASE']['username']
PASSWORD = config['DATABASE']['password']
DSN = config['DATABASE']['dsn']
connectionString = USERNAME+'/'+PASSWORD+'@'+DSN
startmonth = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
STARTMONTH = startmonth.strftime("%Y%m") + '26'
ENDMONTH = datetime.now().strftime("%Y%m") + '25'

def open_file(fileName):
    with open("sql_script/"+fileName, "r") as f:
        data = f.read()
        return data


def save_trx_to_excel():
    try:
        with oracledb.connect(connectionString) as connection:
            print("Connected")

            try:
                getTrxScript = open_file("Trx.sql")
                getTrxScript = getTrxScript.replace(":startMonth26", STARTMONTH)
                getTrxScript = getTrxScript.replace(":endMonth25", ENDMONTH)
                with connection.cursor() as cursor:
                    for result in cursor.execute(getTrxScript):
                        print("data: ", result)
            except oracledb.DatabaseError as e:
                print("Cursor error due to: ", e)
    except oracledb.DatabaseError as e:
        print("Error connecting to Database: ", e)


save_trx_to_excel()
# open_file("Trx.sql")



